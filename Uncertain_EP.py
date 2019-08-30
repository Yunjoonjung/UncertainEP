import numpy as np
import openpyxl

from pyDOE import lhs

from scipy.stats import norm 
from scipy.stats import uniform 
from scipy.stats import triang
from scipy.stats import lognorm

from pyepw.epw import EPW

from SALib.analyze import morris 

from eppy import modeleditor
from eppy.modeleditor import IDF
from eppy.results import readhtml
import pprint

import witheppy
import witheppy.runner

import seaborn as sns
import matplotlib.pyplot as plt

np.set_printoptions(linewidth=2000)
np.set_printoptions(formatter={'float': '{: 0.2f}'.format})

                    
class Uncertain_EP(object):
#-----------------------------------------------------------------------------------------------------------------------------------#     
    def __init__(self, IDF_FileName, epw_FileName, IDD_FileName, outputFile, climate_uncertainty=True, SA_Graph=True, UA_Graph=True):
        self.IDF_FileName = IDF_FileName
        self.epw_FileName = epw_FileName
        self.IDD_FileName = IDD_FileName
        self.outputFile = outputFile
        self.climate_uncertainty = climate_uncertainty
        self.SA_Graph = SA_Graph
        self.UA_Graph = UA_Graph

        # Open .epw file
        epw = EPW()
        epw.read(self.epw_FileName)

        # Save original dry_bulb_temperature and wind_speed
        self.original_EPW = np.zeros((8760,2)) # 0th: dry_bulb_temperature, 1st: wind_speed

        for i, wd in enumerate(epw.weatherdata):
            self.original_EPW[i,0] = wd.dry_bulb_temperature
            self.original_EPW[i,1] = wd.wind_speed

        # Close EPW file
        epw.save(self.epw_FileName)
        
#-----------------------------------------------------------------------------------------------------------------------------------#     
    def EPW_Uncertainty_Propagation(self):
        # if climate uncertainty propagation is needed, this method is executed.

        # Propagated Uncertain Values
        self.propagated_EPW = np.zeros((8760,2)) # 0th: dry_bulb_temperature, 1st: wind_speed

        # Unceratinty propagation
        design_lhs_temp = lhs(1, samples=8760) # 0th: V_met, 1st: alpha, 2nd: delta, 3rd: alpha_met, 4th: delta_met
        design_lhs_wind = lhs(5, samples=8760) # 0th: T_met

        temp_dist = np.zeros((8760, 1))
        wind_dist = np.zeros((8760, 5))

        # Temperature propagation
        temp_dist[:,0] = norm(loc=18, scale=3.4).ppf(design_lhs_temp[:,0]) # T_met
    
        # Wind propagation
        wind_dist[:,0] = lognorm(0.96, 0.23).ppf(design_lhs_wind[:,0]) # V_met 
        wind_dist[:,1] = triang(c=0.48, loc=0.10, scale=0.25).ppf(design_lhs_wind[:,1]) # alpha 
        wind_dist[:,2] = triang(c=0.8, loc=210, scale=200).ppf(design_lhs_wind[:,2]) # delta
        wind_dist[:,3] = triang(c=0.16, loc=0.10, scale=0.25).ppf(design_lhs_wind[:,3]) # alpha_met
        wind_dist[:,4] = triang(c=0.3, loc=210, scale=200).ppf(design_lhs_wind[:,4]) # delta_met
              
        # Temperature Uncertainty Propagation
        L = -0.0065
        H_b = 0
        E = 6356
        E_met = 6356.3
        z = 8
        z_met = 3
        for i in range(8760):
##            self.propagated_EPW[i,0] = (self.original_EPW[i,0] - L*( (E_met/(6356+z_met) - H_b)) + L*( (E/ (E+z) - H_b))) + temp_dist[i,0]
            self.propagated_EPW[i,0] = (self.original_EPW[i,0] - L*( (E_met/(6356+z_met) - H_b)) + L*( (E/ (E+z) - H_b)))

        # Wind Uncertainty Propagation
        z_met = 10
        for i in range(8760):
##            self.propagated_EPW[i,1] = (self.original_EPW[i,1] * ((wind_dist[i,4]/z_met)**wind_dist[i,3]) * ((z/wind_dist[i,2])**wind_dist[i,1])) + wind_dist[i,0]
            self.propagated_EPW[i,1] = (self.original_EPW[i,1] * ((wind_dist[i,4]/z_met)**wind_dist[i,3]) * ((z/wind_dist[i,2])**wind_dist[i,1])) 

        # Open .epw file
        epw = EPW()
        epw.read(self.epw_FileName)
        
        # Assign Propagated values in the .epw file
        for i, wd in enumerate(epw.weatherdata):
            wd.dry_bulb_temperature = self.propagated_EPW[i,0]
            wd.wind_speed = self.propagated_EPW[i,1]
            
        # Close EPW file
##        self.new_file_name = "Climate_Uncertainty_Propagated_TMY.epw"
        epw.save("Climate_Uncertainty_Propagated_TMY.epw")


#-----------------------------------------------------------------------------------------------------------------------------------#     
    def EP_iteration(self, SA_quantified_matrix):  # This is used only for sensitivity analysis
        SA_result_compilation  = np.zeros([SA_quantified_matrix.shape[0]])

        # Read the "Uncertain_EP_Input" excel file
        uncertain_input = openpyxl.load_workbook('Uncertain_EP_Input.xlsx', data_only=False)
        uncertain_input_sheet = uncertain_input['Input']
        
        # Assign the quantified values into idf and epw
        for j, X in enumerate(SA_quantified_matrix):
            # Open .idf file
            IDF.setiddname(self.IDD_FileName)
            instance_idf = IDF(self.IDF_FileName)
            
            loop_count = 0
            for i in range(self.number_of_parameter_uncertain_parameters):
                obj_num = int(uncertain_input_sheet.cell(row=i+2, column=4).value[3]) # int type
                
                if uncertain_input_sheet.cell(row=2, column=2).value == "Material":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Conductivity":
                        instance_idf.idfobjects['Material'][obj_num].Conductivity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Thickness":
                        instance_idf.idfobjects['Material'][obj_num].Thickness = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Density":
                        instance_idf.idfobjects['Material'][obj_num].Density = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Specific_Heat":
                        instance_idf.idfobjects['Material'][obj_num].Specific_Heat = X[loop_count]
                        loop_count += 1
                
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "WindowMaterial_SimpleGlazingSystem":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "UFactor":
                        instance_idf.idfobjects['WindowMaterial_SimpleGlazingSystem'][obj_num].UFactor = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Solar_heat_Gain_Coefficient":
                        instance_idf.idfobjects['WindowMaterial_SimpleGlazingSystem'][obj_num].Solar_heat_Gain_Coefficient = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "People":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "People_per_Zone_Floor_Area":
                        instance_idf.idfobjects['People'][obj_num].People_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Zone_Floor_Area_per_Person":
                        instance_idf.idfobjects['People'][obj_num].Zone_Floor_Area_per_Person = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Lights":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Lighting_Level":
                        instance_idf.idfobjects['Lights'][obj_num].Lighting_Level = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Zone_Floor_Area":
                        instance_idf.idfobjects['Lights'][obj_num].Watts_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Person":
                        instance_idf.idfobjects[''][obj_num].Watts_per_Person = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "ElectricEquipment":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Design_Level":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Design_Level = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Zone_Floor_Area":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Watts_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Person":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Watts_per_Person = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "ZoneInfiltration_DesignFlowRate":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Design_Flow_Rate":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Design_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Flow_per_Zone_Floor_Area":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Flow_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Flow_per_Exterior_Surface_Area":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Flow_per_Exterior_Surface_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Air_Changes_per_Hour":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Air_Changes_per_Hour = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Effective_Leakage_Area":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Effective_Leakage_Area = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "AirflowNetwork_MultiZone_WindPressureCoefficientValues":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_1":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_1 = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_2":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_2= X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_3":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_3 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_4":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_4 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_5":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_5 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_6":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_6 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_7":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_7 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_8":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_8 = X[loop_count]   

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_9":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_9 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_10":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_10 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_11":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_11 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_12":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_12 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_13":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_13 = X[loop_count]
                        loop_count += 1
             
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Fan_VariableVolume":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Fan_Total_Efficiency":
                        instance_idf.idfobjects['Fan_VariableVolume'][obj_num].Fan_Total_Efficiency = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Pressure_Rise":
                        instance_idf.idfobjects['Fan_VariableVolume'][obj_num].Pressure_Rise = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Cooling_DX_SingleSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Cooling_DX_TwoSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "CoilPerformance_DX_Cooling":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Heating_DX_SingleSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Heating_Capacity":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Gross_Rated_Heating_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Heating_COP":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Gross_Rated_Heating_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
        
            # Run EnergyPlus
            instance_idf.save()
            # Scenario Uncertainty Propagation
            if self.climate_uncertainty == True:
                self.EPW_Uncertainty_Propagation()
                temporary_weather_file_name = r"Climate_Uncertainty_Propagated_TMY.epw"
            
                idf = IDF(self.IDF_FileName,temporary_weather_file_name)
                witheppy.runner.eplaunch_run(idf)
            else:
                idf = IDF(self.IDF_FileName, self.epw_FileName)
                witheppy.runner.eplaunch_run(idf)
                
            # Collect the result
            fname = self.outputFile # the html file you want to read
            filehandle = open(fname, 'r').read() # get a file handle to the html file
            htables = readhtml.titletable(filehandle) # reads the tables with their titles
            firstitem = htables[0]
            seconditem = htables[2]

            pp = pprint.PrettyPrinter()
##                pp.pprint(firstitem)
##                print(firstitem[1][1][1])
            total_site_energy = firstitem[1][1][1]
##                pp.pprint(seconditem)
            total_conditioned_area = seconditem[1][2][1]

            
            total_site_energy = (firstitem[1][1][1])*277.778/total_conditioned_area
            SA_result_compilation[j] = total_site_energy
        
        uncertain_input.save('Uncertain_EP_Input.xlsx') # close Excel
        
        return SA_result_compilation
    

#-----------------------------------------------------------------------------------------------------------------------------------#     
    def Uncertain_Quantification(self):

        # Read the "Uncertain_EP_Input" excel file
        uncertain_input = openpyxl.load_workbook('Uncertain_EP_Input.xlsx', data_only=False)
        uncertain_input_sheet = uncertain_input['Input']

        self.number_of_parameter_uncertain_parameters = int(uncertain_input_sheet.cell(row=2, column=14).value)
        self.distribution_repository = np.zeros((self.number_of_samples, self.number_of_parameter_uncertain_parameters))
        
        # Uncertainty quantification (only parameter uncertainty)
        design_lhs = lhs(self.number_of_parameter_uncertain_parameters, samples=self.number_of_samples)

        # Open .idf file
        IDF.setiddname(self.IDD_FileName)
        instance_idf = IDF(self.IDF_FileName)

        value_in_idf = 0 # For saving values in an idf instance
        name_repository = [] # For saving parameters (purpose: sensivitiy analysis drawing graph)
        
        for i in range(0,self.number_of_parameter_uncertain_parameters):
            replace_to_EnergyPlus_format = str(uncertain_input_sheet.cell(row=i+2, column=2).value.replace("_", ":")) # Convert from _ to :
            obj_num = int(uncertain_input_sheet.cell(row=i+2, column=4).value[3]) # find object number
            field_name= str(uncertain_input_sheet.cell(row=i+2, column=3).value) # find field name
            class_in_idf_instance = instance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
 
            name_in_idf_instance = instance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1].Name
            sample_string = name_in_idf_instance +"_IN_"+ str(uncertain_input_sheet.cell(row=i+2, column=3).value)
            name_repository.append(sample_string) # This is only for sensitivity analysis
                
#-----------#-------------------------------------------------------------------------------------------------------------------------#        
            # 1. Normal Relative distribution
            if uncertain_input_sheet.cell(row=i+2, column=5).value == "NormalRelative":
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        value_in_idf = class_in_idf_instance.Conductivity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        value_in_idf = class_in_idf_instance.Thickness
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        value_in_idf = class_in_idf_instance.Density
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        value_in_idf = class_in_idf_instance.Specific_Heat
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        value_in_idf = class_in_idf_instance.UFactor
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        value_in_idf =  class_in_idf_instance.Solar_heat_Gain_Coefficient
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                        
                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        class_in_idf_instance = class_in_idf_instance.People_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        class_in_idf_instance = class_in_idf_instance.Zone_Floor_Area_per_Person
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        value_in_idf = class_in_idf_instance.Lighting_Level
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        value_in_idf = class_in_idf_instance.Design_Level
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                            value_in_idf = class_in_idf_instance.Design_Flow_Rate
                            self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Exterior_Surface_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Air_Changes_per_Hour
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Person
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_Air_Changes_per_Hour
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        value_in_idf = class_in_idf_instance.Effective_Leakage_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_1
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_2
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_3
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_4
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_5
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_6
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_7
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_8
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_9
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_10
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_11
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_12
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_13
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Maximum_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        value_in_idf = class_in_idf_instance.Fan_Total_Efficiency
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                                   
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance_DX_Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_Capacity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_COP
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                
#-----------#-------------------------------------------------------------------------------------------------------------------------#        
            # 2. UNIFORM Relative Distribution 
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "UniformRelative":                
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        value_in_idf = class_in_idf_instance.Conductivity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        value_in_idf = class_in_idf_instance.Thickness
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        value_in_idf = class_in_idf_instance.Density
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        value_in_idf = class_in_idf_instance.Specific_Heat
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        value_in_idf = class_in_idf_instance.UFactor
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        value_in_idf =  class_in_idf_instance.Solar_heat_Gain_Coefficient
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        class_in_idf_instance = class_in_idf_instance.People_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        class_in_idf_instance = class_in_idf_instance.Zone_Floor_Area_per_Person
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
            
                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        value_in_idf = class_in_idf_instance.Lighting_Level
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        value_in_idf = class_in_idf_instance.Design_Level
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Design_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Exterior_Surface_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Air_Changes_per_Hour
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Person
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_Air_Changes_per_Hour
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        value_in_idf = class_in_idf_instance.Effective_Leakage_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_1
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_2
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_3
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_4
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_5
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_6
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_7
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_8
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_9
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_10
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_11
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_12
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_13
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Maximum_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        value_in_idf = class_in_idf_instance.Fan_Total_Efficiency
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                                       
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance:DX:Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_Capacity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_COP
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

#-----------#-------------------------------------------------------------------------------------------------------------------------#        
            # 3. Triangle Relative Distribution  
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "TriangleRelative":
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
            
                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                                       
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance:DX:Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
       
#-----------#-------------------------------------------------------------------------------------------------------------------------#        
            # 4. Normal Absolute Distribution
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "NormalAbsolute":            
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                        
                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                                   
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance_DX_Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                        
#-----------#-------------------------------------------------------------------------------------------------------------------------#        
            # 5. Uniform Absolute Distribution
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "UniformAbsolute":                
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
            
                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                                       
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance:DX:Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

#-----------#-------------------------------------------------------------------------------------------------------------------------#        
            # 6. Triangle Absolute Distribution
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "TriangleAbsolute":                
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
            
                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                                       
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance:DX:Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

#-----------#-------------------------------------------------------------------------------------------------------------------------#        
            # 7. LogNormal Absolute Distribution
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "LogNormalAbsolute":
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                        
                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                                   
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance_DX_Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        self.distribution_repository[:,i] = lognorm(uncertain_input_sheet.cell(row=i+2, column=6).value, uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        self.distribution_repository[:,i] = norm(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])


        instance_idf.save() # close idf 
        uncertain_input.save('Uncertain_EP_Input.xlsx') # close Excel

        print(self.distribution_repository)
        
        return self.distribution_repository, name_repository


#-----------------------------------------------------------------------------------------------------------------------------------#
    def SA(self, number_of_SA_samples=1000):
        self.number_of_samples = number_of_SA_samples
        
        # Read the "Uncertain_EP_Input" excel file
        uncertain_input = openpyxl.load_workbook('Uncertain_EP_Input.xlsx', data_only=False)
        uncertain_input_sheet = uncertain_input['Input']

        self.number_of_parameter_uncertain_parameters = int(uncertain_input_sheet.cell(row=2, column=14).value)
        self.distribution_repository = np.zeros((self.number_of_samples, self.number_of_parameter_uncertain_parameters))
        
        # Define the variables for Morris sensitivity analysis
        self.energyPlus_input_setup = {'num_vars': self.number_of_parameter_uncertain_parameters}
        
        # Call the "Uncertain_Quantification" method
        quantified_repository, name_repository = self.Uncertain_Quantification()
        
        self.energyPlus_input_setup.update({'names': name_repository})
        print(self.energyPlus_input_setup)   
 
        # Conduct Morris Method
        Y = self.EP_iteration(quantified_repository)
        Si = morris.analyze(self.energyPlus_input_setup, quantified_repository, Y, conf_level=0.95, print_to_console=True, num_levels=4)
        print(Si)

        uncertain_input.save('Uncertain_EP_Input.xlsx') # close Excel

        # Visualize the morris method result
        if self.SA_Graph  == True:
            pass
        #-------------------------------------------------#
        #-------------------------------------------------#
        #-------------------------------------------------#
        #-------------------------------------------------#
        #-------------------------------------------------#


        
#-----------------------------------------------------------------------------------------------------------------------------------#
    def UA(self, number_of_UA_samples=1000, only_idf_instances_generation=True):
        self.number_of_samples = number_of_UA_samples

        # Read the "Uncertain_EP_Input" excel file
        uncertain_input = openpyxl.load_workbook('Uncertain_EP_Input.xlsx', data_only=False)
        uncertain_input_sheet = uncertain_input['Input']

        # Call the "Uncertain_Quantification" method to quantify uncertain parameters
        quantified_repository, name_repository = self.Uncertain_Quantification()
        self.UA_output_result_repository = np.zeros([quantified_repository.shape[0]]) # Fior visualization
          
        # ASSIGN the quantified values into idf and epw
        for j, X in enumerate(self.distribution_repository):
            # Open .idf file
            IDF.setiddname(self.IDD_FileName)
            instance_idf = IDF(self.IDF_FileName)
            
            # Assign the quantified values           
            loop_count = 0
            for i in range(self.number_of_parameter_uncertain_parameters):
                obj_num = int(uncertain_input_sheet.cell(row=i+2, column=4).value[3])-1
                if uncertain_input_sheet.cell(row=2, column=2).value == "Material":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Conductivity":
                        instance_idf.idfobjects['Material'][obj_num].Conductivity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Thickness":
                        instance_idf.idfobjects['Material'][obj_num].Thickness = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Density":
                        instance_idf.idfobjects['Material'][obj_num].Density = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Specific_Heat":
                        instance_idf.idfobjects['Material'][obj_num].Specific_Heat = X[loop_count]
                        loop_count += 1                
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "People":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "People_per_Zone_Floor_Area":
                        instance_idf.idfobjects['People'][obj_num].People_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Zone_Floor_Area_per_Person":
                        instance_idf.idfobjects['People'][obj_num].Zone_Floor_Area_per_Person = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Lights":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Lighting_Level":
                        instance_idf.idfobjects['Lights'][obj_num].Lighting_Level = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Zone_Floor_Area":
                        instance_idf.idfobjects['Lights'][obj_num].Watts_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Person":
                        instance_idf.idfobjects[''][obj_num].Watts_per_Person = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "ElectricEquipment":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Design_Level":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Design_Level = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Zone_Floor_Area":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Watts_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Person":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Watts_per_Person = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "ZoneInfiltration_DesignFlowRate":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Design_Flow_Rate":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Design_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Flow_per_Zone_Floor_Area":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Flow_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Flow_per_Exterior_Surface_Area":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Flow_per_Exterior_Surface_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Air_Changes_per_Hour":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Air_Changes_per_Hour = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Effective_Leakage_Area":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Effective_Leakage_Area = X[loop_count]
                        loop_count += 1
                    

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "AirflowNetwork_MultiZone_WindPressureCoefficientValues":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_1":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_1 = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_2":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_2= X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_3":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_3 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_4":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_4 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_5":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_5 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_6":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_6 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_7":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_7 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_8":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_8 = X[loop_count]   

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_9":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_9 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_10":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_10 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_11":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_11 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_12":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_12 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_13":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_13 = X[loop_count]
                        loop_count += 1
             
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Fan_VariableVolume":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Fan_Total_Efficiency":
                        instance_idf.idfobjects['Fan_VariableVolume'][obj_num].Fan_Total_Efficiency = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Pressure_Rise":
                        instance_idf.idfobjects['Fan_VariableVolume'][obj_num].Pressure_Rise = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Cooling_DX_SingleSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Cooling_DX_TwoSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "CoilPerformance_DX_Cooling":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Heating_DX_SingleSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Heating_Capacity":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Gross_Rated_Heating_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Heating_COP":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Gross_Rated_Heating_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1      

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Thickness":
                        instance_idf.idfobjects['Material'][obj_num].Thickness = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Density":
                        instance_idf.idfobjects['Material'][obj_num].Density = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Specific_Heat":
                        instance_idf.idfobjects['Material'][obj_num].Specific_Heat = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "WindowMaterial_SimpleGlazingSystem":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "UFactor":
                        instance_idf.idfobjects['WindowMaterial_SimpleGlazingSystem'][obj_num].UFactor = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Solar_heat_Gain_Coefficient":
                        instance_idf.idfobjects['WindowMaterial_SimpleGlazingSystem'][obj_num].Solar_heat_Gain_Coefficient = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "People":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "People_per_Zone_Floor_Area":
                        instance_idf.idfobjects['People'][obj_num].People_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Zone_Floor_Area_per_Person":
                        instance_idf.idfobjects['People'][obj_num].Zone_Floor_Area_per_Person = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Lights":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Lighting_Level":
                        instance_idf.idfobjects['Lights'][obj_num].Lighting_Level = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Zone_Floor_Area":
                        instance_idf.idfobjects['Lights'][obj_num].Watts_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Person":
                        instance_idf.idfobjects[''][obj_num].Watts_per_Person = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "ElectricEquipment":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Design_Level":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Design_Level = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Zone_Floor_Area":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Watts_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Person":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Watts_per_Person = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "ZoneInfiltration_DesignFlowRate":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Design_Flow_Rate":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Design_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Flow_per_Zone_Floor_Area":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Flow_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Flow_per_Exterior_Surface_Area":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Flow_per_Exterior_Surface_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Air_Changes_per_Hour":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Air_Changes_per_Hour = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Effective_Leakage_Area":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Effective_Leakage_Area = X[loop_count]
                        loop_count += 1
                    

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "AirflowNetwork_MultiZone_WindPressureCoefficientValues":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_1":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_1 = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_2":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_2= X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_3":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_3 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_4":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_4 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_5":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_5 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_6":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_6 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_7":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_7 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_8":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_8 = X[loop_count]   

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_9":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_9 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_10":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_10 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_11":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_11 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_12":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_12 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_13":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_13 = X[loop_count]
                        loop_count += 1
             
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Fan_VariableVolume":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Fan_Total_Efficiency":
                        instance_idf.idfobjects['Fan_VariableVolume'][obj_num].Fan_Total_Efficiency = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Pressure_Rise":
                        instance_idf.idfobjects['Fan_VariableVolume'][obj_num].Pressure_Rise = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Cooling_DX_SingleSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Cooling_DX_TwoSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "CoilPerformance_DX_Cooling":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Heating_DX_SingleSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Heating_Capacity":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Gross_Rated_Heating_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Heating_COP":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Gross_Rated_Heating_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
             
            #-----------------------------------------------------------------------------------------------------------------------------------------------------#
            if only_idf_instances_generation == True: # Only Quantification and save-as idf instances case
                idf_instance_name = "idf_instance_" + str(j+1)
                print(idf_instance_name+" was generated.")
                idf_save_path = "./Output/"+idf_instance_name+".idf"
                instance_idf.save(idf_save_path)

            elif only_idf_instances_generation == False: # Simulate EnergyPlus case
                # Run EnergyPlus
                instance_idf.save()
                
                # Scenario Uncertainty Propagation
                if self.climate_uncertainty == True:
                    self.EPW_Uncertainty_Propagation()
                    temporary_weather_file_name =  r"Climate_Uncertainty_Propagated_TMY.epw"
            
                    idf = IDF(self.IDF_FileName,temporary_weather_file_name)
                    witheppy.runner.eplaunch_run(idf)
                else:   
                    idf = IDF(self.IDF_FileName, self.epw_FileName)
                    witheppy.runner.eplaunch_run(idf)
                    
                # Collect the result
                fname = self.outputFile # the html file you want to read
                filehandle = open(fname, 'r').read() # get a file handle to the html file
                htables = readhtml.titletable(filehandle) # reads the tables with their titles
                firstitem = htables[0]
                seconditem = htables[2]

                pp = pprint.PrettyPrinter()
##                pp.pprint(firstitem)
##                print(firstitem[1][1][1])
                total_site_energy = firstitem[1][1][1]
##                pp.pprint(seconditem)
                total_conditioned_area = seconditem[1][2][1]
              
                self.UA_output_result_repository[j] = total_site_energy*277.778/total_conditioned_area

            
        uncertain_input.save('Uncertain_EP_Input.xlsx') # close Excel
        
        if self.UA_Graph == True:
            graph_for_UA = sns.distplot(self.UA_output_result_repository[:], kde=False, rug=False)
            graph_for_UA.set(xlabel="Delivered Energy [kWh/m2/yr]", ylabel = "Frequency")
            plt.show()

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
        
