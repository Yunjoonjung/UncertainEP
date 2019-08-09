import numpy as np

import openpyxl

from SALib.analyze import morris 

from scipy.stats import norm 
from scipy.stats import uniform 
from scipy.stats import triang

from pyepw.epw import EPW

from pyDOE import lhs

import witheppy
import witheppy.runner

from eppy import modeleditor
from eppy.modeleditor import IDF
from eppy.results import readhtml
import pprint

import seaborn as sns
import matplotlib.pyplot as plt

np.set_printoptions(linewidth=2000)

                    
class Uncertain_EP(object):
    
    def __init__(self, IDF_FileName, epw_FileName, IDD_FileName, SA_Graph=True, UA_Graph=True):
        self.epw_FileName = epw_FileName
        self.IDF_FileName = IDF_FileName
        self.IDD_FileName = IDD_FileName
        self.SA_Graph = SA_Graph
        self.UA_Graph = UA_Graph
        
        
    def EP_iteration(self, SA_quantified_matrix): 
        SA_result_compilation  = np.zeros([SA_quantified_matrix.shape[0]])

        # Read the "Uncertain_EP_Input" excel file
        uncertain_input = openpyxl.load_workbook('Uncertain_EP_Input2.xlsx', data_only=False)
        uncertain_input_sheet = uncertain_input['Input']
        
        # Assign the quantified values into idf and epw
        for j, X in enumerate(SA_quantified_matrix):
            # Open .idf file
            IDF.setiddname(self.IDD_FileName)
            instance_idf = IDF(self.IDF_FileName)
        
            loop_count = 0
            for i in range(self.number_of_parameter_uncertain_parameters):
                obj_num = int(uncertain_input_sheet.cell(row=i+2, column=4).value[3]) # int type
                if uncertain_input_sheet.cell(row=2, column=2).value == "Material":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Conductivity":
                        instance_idf.idfobjects['Material'][obj_num].Conductivity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Thickness":
                        instance_idf.idfobjects['Material'][obj_num].Thickness = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Density":
                        instance_idf.idfobjects['Material'][obj_num].Density = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Specific_Heat":
                        instance_idf.idfobjects['Material'][obj_num].Specific_Heat = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "WindowMaterial_SimpleGlazingSystem":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "UFactor":
                        instance_idf.idfobjects['WindowMaterial_SimpleGlazingSystem'][obj_num].UFactor = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Solar_heat_Gain_Coefficient":
                        instance_idf.idfobjects['WindowMaterial_SimpleGlazingSystem'][obj_num].Solar_heat_Gain_Coefficient = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "People":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "People_per_Zone_Floor_Area":
                        instance_idf.idfobjects['People'][obj_num].People_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Zone_Floor_Area_per_Person":
                        instance_idf.idfobjects['People'][obj_num].Zone_Floor_Area_per_Person = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Lights":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Lighting_Level":
                        instance_idf.idfobjects['Lights'][obj_num].Lighting_Level = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Zone_Floor_Area":
                        instance_idf.idfobjects['Lights'][obj_num].Watts_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Person":
                        instance_idf.idfobjects[''][obj_num].Watts_per_Person = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "ElectricEquipment":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Design_Level":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Design_Level = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Zone_Floor_Area":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Watts_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Watts_per_Person":
                        instance_idf.idfobjects['ElectricEquipment'][obj_num].Watts_per_Person = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "ZoneInfiltration_DesignFlowRate":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Design_Flow_Rate":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Design_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Flow_per_Zone_Floor_Area":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Flow_per_Zone_Floor_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Flow_per_Exterior_Surface_Area":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Flow_per_Exterior_Surface_Area = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Air_Changes_per_Hour":
                        instance_idf.idfobjects['ZoneInfiltration_DesignFlowRate'][obj_num].Air_Changes_per_Hour = X[loop_count]
                        loop_count += 1


                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Effective_Leakage_Area":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Effective_Leakage_Area = X[loop_count]
                        loop_count += 1
                    

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "AirflowNetwork_MultiZone_WindPressureCoefficientValues":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_1":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_1 = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_2":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_2= X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_3":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_3 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_4":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_4 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_5":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_5 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_6":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_6 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_7":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_7 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_8":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_8 = X[loop_count]   

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_9":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_9 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_10":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_10 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_11":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_11 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_12":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_12 = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Wind_Pressure_Coefficient_Value_13":
                        instance_idf.idfobjects['AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea'][obj_num].Wind_Pressure_Coefficient_Value_13 = X[loop_count]
                        loop_count += 1
             
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Fan_VariableVolume":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Fan_Total_Efficiency":
                        instance_idf.idfobjects['Fan_VariableVolume'][obj_num].Fan_Total_Efficiency = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Pressure_Rise":
                        instance_idf.idfobjects['Fan_VariableVolume'][obj_num].Pressure_Rise = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Cooling_DX_SingleSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Cooling_DX_SingleSpeed'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                        
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Cooling_DX_TwoSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "High_Speed_Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Cooling_DX_TwoSpeed'][obj_num].High_Speed_Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "CoilPerformance_DX_Cooling":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Total_Cooling_Capacity":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Total_Cooling_Capacity = X[loop_count]
                        loop_count += 1
                    
                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Sensible_Heat_Ratio":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Sensible_Heat_Ratio = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Cooling_COP":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Gross_Rated_Cooling_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['CoilPerformance_DX_Cooling'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1

                elif uncertain_input_sheet.cell(row=i+2, column=2).value == "Coil_Heating_DX_SingleSpeed":
                    if uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Heating_Capacity":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Gross_Rated_Heating_Capacity = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Gross_Rated_Heating_COP":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Gross_Rated_Heating_COP = X[loop_count]
                        loop_count += 1

                    elif uncertain_input_sheet.cell(row=i+2, column=3).value == "Rated_Air_Flow_Rate":
                        instance_idf.idfobjects['Coil_Heating_DX_SingleSpeed'][obj_num].Rated_Air_Flow_Rate = X[loop_count]
                        loop_count += 1
                    
        
            # Run
            instance_idf.save()
            idf = IDF(self.IDF_FileName, self.epw_FileName)
##            idf.run()
            witheppy.runner.eplaunch_run(idf)
                
            # Collect the result
            fname = "New_TGSTable.htm" # the html file you want to read
            filehandle = open(fname, 'r').read() # get a file handle to the html file
            htables = readhtml.titletable(filehandle) # reads the tables with their titles
            firstitem = htables[0]
##            print(firstitem)

            pp = pprint.PrettyPrinter()
            pp.pprint(firstitem)
            print(firstitem[1][1][1])
            total_site_energy = firstitem[1][1][1]
            SA_result_compilation[j] = total_site_energy
            
##        a = sns.distplot(output[:])
##        plt.show()
        
        uncertain_input.save('Uncertain_EP_Input2.xlsx') # close Excel
        return SA_result_compilation 

        
    def SA(self, number_of_samples=11):
        self.number_of_samples = number_of_samples
        self.uncertain_parameter_repository = []
        value_in_idf = 0 # For saving values in an idf instance

        # Read the "Uncertain_EP_Input" excel file
        uncertain_input = openpyxl.load_workbook('Uncertain_EP_Input2.xlsx', data_only=True)
        uncertain_input_sheet = uncertain_input['Input']

        self.number_of_parameter_uncertain_parameters = uncertain_input_sheet.cell(row=2, column=14).value
        
        self.distribution_repository = np.zeros((self.number_of_samples, self.number_of_parameter_uncertain_parameters))
      
        # Open .idf file
        IDF.setiddname(self.IDD_FileName)
        instance_idf = IDF(self.IDF_FileName)
        
        # Define the variables for Morris sensitivity analysis
        self.energyPlus_input_setup = {'num_vars': self.number_of_parameter_uncertain_parameters}

        name_repository = []
        for i in range(0,self.number_of_parameter_uncertain_parameters):
            replace_to_EnergyPlus_format = str(uncertain_input_sheet.cell(row=i+2, column=2).value.replace("_", ":"))
            replace_to_EnergyPlus_format =replace_to_EnergyPlus_format.upper()
            
            obj_num = int(uncertain_input_sheet.cell(row=i+2, column=4).value[3])
            
            name_in_idf_instance = instance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1].Name

            sample_string = name_in_idf_instance +"_IN_"+ str(uncertain_input_sheet.cell(row=i+2, column=3).value)
            name_repository.append(sample_string)

        self.energyPlus_input_setup.update({'names': name_repository})

                
        # Uncertainty quantification (only parameter uncertainty)
        design_lhs = lhs(self.number_of_parameter_uncertain_parameters, samples=self.number_of_samples)
        
        for i in range(0,self.number_of_parameter_uncertain_parameters):
            replace_to_EnergyPlus_format = str(uncertain_input_sheet.cell(row=i+2, column=2).value.replace("_", ":")) # Convert from _ to :
##            print(replace_to_EnergyPlus_format)
            obj_num = int(uncertain_input_sheet.cell(row=i+2, column=4).value[3]) # find object number
            field_name= str(uncertain_input_sheet.cell(row=i+2, column=3).value) # find field name           
            class_in_idf_instance = instance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
##            print(class_in_idf_instance)
                
            if uncertain_input_sheet.cell(row=i+2, column=5).value == "NormalRelative":
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        value_in_idf = class_in_idf_instance.Conductivity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        value_in_idf = class_in_idf_instance.Thickness
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        value_in_idf = class_in_idf_instance.Density
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        value_in_idf = class_in_idf_instance.Specific_Heat
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        value_in_idf = class_in_idf_instance.UFactor
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        value_in_idf =  class_in_idf_instance.Solar_heat_Gain_Coefficient
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        class_in_idf_instance = class_in_idf_instance.People_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        class_in_idf_instance = class_in_idf_instance.Zone_Floor_Area_per_Person
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        value_in_idf = class_in_idf_instance.Lighting_Level
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        value_in_idf = class_in_idf_instance.Design_Level
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                            value_in_idf = class_in_idf_instance.Design_Flow_Rate
                            self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Exterior_Surface_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Air_Changes_per_Hour
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Person
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_Air_Changes_per_Hour
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork_MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        value_in_idf = class_in_idf_instance.Effective_Leakage_Area
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_1
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_2
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_3
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_4
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_5
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_6
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_7
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_8
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_9
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_10
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_11
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_12
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_13
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Maximum_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        value_in_idf = class_in_idf_instance.Fan_Total_Efficiency
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                                   
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance_DX_Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_Capacity
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_COP
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = norm(loc=value_in_idf, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])


            # UNIFORM Distribution case
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "UniformRelative":                
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        value_in_idf = class_in_idf_instance.Conductivity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        value_in_idf = class_in_idf_instance.Thickness
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        value_in_idf = class_in_idf_instance.Density
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        value_in_idf = class_in_idf_instance.Specific_Heat
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        value_in_idf = class_in_idf_instance.UFactor
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        value_in_idf =  class_in_idf_instance.Solar_heat_Gain_Coefficient
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        class_in_idf_instance = class_in_idf_instance.People_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        class_in_idf_instance = class_in_idf_instance.Zone_Floor_Area_per_Person
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
            
                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        value_in_idf = class_in_idf_instance.Lighting_Level
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        value_in_idf = class_in_idf_instance.Design_Level
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Design_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Exterior_Surface_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Air_Changes_per_Hour
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Person
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_Air_Changes_per_Hour
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        value_in_idf = class_in_idf_instance.Effective_Leakage_Area
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_1
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_2
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_3
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_4
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_5
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_6
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_7
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_8
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_9
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_10
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_11
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_12
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_13
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Maximum_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        value_in_idf = class_in_idf_instance.Fan_Total_Efficiency
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                                       
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance:DX:Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_Capacity
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_COP
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])



            # Triangle       
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "TriangleRelative":
                if replace_to_EnergyPlus_format == 'Material':
                    if field_name == 'Conductivity':
                        value_in_idf = class_in_idf_instance.Conductivity
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                        
                    elif field_name == 'Thickness':
                        value_in_idf = class_in_idf_instance.Thickness
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Density':
                        value_in_idf = class_in_idf_instance.Density
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Specific_Heat':
                        value_in_idf = class_in_idf_instance.Specific_Heat
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'WindowMaterial:SimpleGlazingSystem':
                    if field_name == 'UFactor':
                        value_in_idf = class_in_idf_instance.UFactor
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        value_in_idf =  class_in_idf_instance.Solar_heat_Gain_Coefficient
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'People':
                    if field_name == 'People_per_Zone_Floor_Area':
                        class_in_idf_instance = class_in_idf_instance.People_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        class_in_idf_instance = class_in_idf_instance.Zone_Floor_Area_per_Person
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
            
                elif replace_to_EnergyPlus_format == 'Lights':
                    if field_name == 'Lighting_Level':
                        value_in_idf = class_in_idf_instance.Lighting_Level
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    if field_name == 'Design_Level':
                        value_in_idf = class_in_idf_instance.Design_Level
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Watts_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Watts_per_Person':
                        value_in_idf = class_in_idf_instance.Watts_per_Person
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration:DesignFlowRate':
                    if field_name == 'Design_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Design_Flow_Rate
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        value_in_idf = class_in_idf_instance.Flow_per_Exterior_Surface_Area
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Air_Changes_per_Hour
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                
                elif replace_to_EnergyPlus_format == 'DesignSpecification:OutdoorAir':
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Person
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone_Floor_Area
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_per_Zone
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        value_in_idf = class_in_idf_instance.Outdoor_Air_Flow_Air_Changes_per_Hour
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'AirflowNetwork:MultiZone:Surface:EffectiveLeakageArea':
                    if field_name == 'Effective_Leakage_Area':
                        value_in_idf = class_in_idf_instance.Effective_Leakage_Area
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_1
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_2
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_3
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_4
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_5
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_6
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_7
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_8
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_9
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_10
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_11
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_12
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        value_in_idf = class_in_idf_instance.Wind_Pressure_Coefficient_Value_13
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan:ConstantVolume':
                    if field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Maximum_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Maximum_Flow_Rate
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Fan_VariableVolume':
                    if field_name == 'Fan_Total_Efficiency':
                        value_in_idf = class_in_idf_instance.Fan_Total_Efficiency
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Pressure_Rise':
                        value_in_idf = class_in_idf_instance.Pressure_Rise
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
                                       
                elif replace_to_EnergyPlus_format == 'Coil:Cooling:DX:TwoSpeed':
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.High_Speed_Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.High_Speed_Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'CoilPerformance:DX:Cooling':
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Total_Cooling_Capacity
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Cooling_COP
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                elif replace_to_EnergyPlus_format == 'Coil:Heating:DX:SingleSpeed':
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_Capacity
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Gross_Rated_Heating_COP':
                        value_in_idf = class_in_idf_instance.Gross_Rated_Heating_COP
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value_in_idf = class_in_idf_instance.Rated_Air_Flow_Rate
                        self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])
       
        
        instance_idf.save() # close idf 
        uncertain_input.save('Uncertain_EP_Input2.xlsx') # close Excel

##        print(self.distribution_repository)
 
        # Conduct Morris Method
        Y = self.EP_iteration(self.distribution_repository)
        Si = morris.analyze(self.energyPlus_input_setup, self.distribution_repository, Y, conf_level=0.95, print_to_console=True, num_levels=4)


        # Visualize the morris method result



                

testing = Uncertain_EP('C:\\Users\\Su Jeong Lee\\Dropbox (GaTech)\\Uncertain_EP\\New_TGS.idf', 'C:\\Users\\Su Jeong Lee\\Dropbox (GaTech)\\Uncertain_EP\\Atlanta.epw', 'C:\\Users\\Su Jeong Lee\\Dropbox (GaTech)\\Uncertain_EP\\Energy+.idd')
testing.SA()
