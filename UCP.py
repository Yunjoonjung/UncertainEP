import numpy as np

import openpyxl

from scipy.stats import norm 
from scipy.stats import uniform 
from scipy.stats import triang

from pyepw.epw import EPW

from pyDOE import lhs

from eppy import modeleditor
from eppy.modeleditor import IDF
from eppy.results import readhtml 
import pprint

import seaborn as sns
import matplotlib.pyplot as plt


class Uncertain_EP(object):
    
    def __init__(self, IDF_FileName, epw_FileName, IDD_FileName, IDD_FileNameDraw_SA_Result=True, Draw_UA_Result=True):
        self.epw_FileName = epw_FileName
        self.IDF_FileName = IDF_FileName
        self.IDD_FileName = IDD_FileName
        self.IDD_FileNameDraw_SA_Result = IDD_FileNameDraw_SA_Result
        self.Draw_UA_Result = Draw_UA_Result
        
        
    def EP_iteration(self, SA_quantified_matrix): # Method for EnergyPlus iteration

        # Read the "Uncertain_EP_Input" excel file
        uncertain_input = openpyxl.load_workbook('Uncertain_EP_Input.xlsx', data_only=False)
        uncertain_input_sheet = uncertain_input['Input']

   
        # Assign the quantified values into idf and epw
        for i, X in enumerate(SA_quantified_matrix):
        
            # Open .idf file
            IDF.setiddname(self.IDD_FileName)
            incetance_idf = IDF(self.IDF_FileName)

            # Assign quantified values in the idf instace
            for j in range(0,int(uncertain_input_sheet.cell(row=1, column=14).value)):
                replace_to_EnergyPlus_format = str(uncertain_input_sheet.cell(row=i[0]+2, column=2).value.replace("_", ":"))
                obj_num = int(uncertain_input_sheet.cell(row=i[0]+2, column=4).value[3])
                name_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1].uncertain_input_sheet.cell(row=i[0]+2, column=3).value

                print(name_in_idf_instance)
                                                                             
                name_in_idf_instance = X[j] #############
                   


            # Run
            incetance_idf.save()
            idf = IDF(self.IDF_FileName, self.epw_FileName)
            idf.run(readvars=True)
            
            
            # Collect the result
            fname = "eplustbl.htm" # the html file you want to read
            filehandle = open(fname, 'r').read() # get a file handle to the html file
            htables = readhtml.titletable(filehandle) # reads the tables with their titles
            firstitem = htables[0]

            pp = pprint.PrettyPrinter()
            pp.pprint(firstitem)
            print(firstitem[1][1][1])
            total_site_energy = firstitem[1][1][1]
            output[i] = total_site_energy

        a = sns.distplot(output[:])
        plt.show()
        
        incetance_idf.save() # close idf 
        uncertain_input.save('Uncertain_EP_Input.xlsx') # close Excel
        return output

        
    def SA(self, number_of_samples=20):
        self.number_of_samples = number_of_samples

        # Read the "Uncertain_EP_Input" excel file
        uncertain_input = openpyxl.load_workbook('Uncertain_EP_Input.xlsx', data_only=True)
        uncertain_input_sheet = uncertain_input['Input']

        self.number_of_parameter_uncertain_parameters = uncertain_input_sheet.cell(row=2, column=14).value
        
        self.distribution_repository = np.zeros((self.number_of_samples, self.number_of_parameter_uncertain_parameters))
      
        # Open .idf file
        IDF.setiddname(self.IDD_FileName)
        incetance_idf = IDF(self.IDF_FileName)
        
        # Define the variables for Morris sensitivity analysis
        self.energyPlus_input_setup = {'num_vars': self.number_of_parameter_uncertain_parameters}

        name_repository = []
        for i in range(0,self.number_of_parameter_uncertain_parameters):
            replace_to_EnergyPlus_format = str(uncertain_input_sheet.cell(row=i+2, column=2).value.replace("_", ":"))
            replace_to_EnergyPlus_format =replace_to_EnergyPlus_format.upper()
            
            obj_num = int(uncertain_input_sheet.cell(row=i+2, column=4).value[3])
            
            name_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1].Name

            sample_string = name_in_idf_instance +"_IN_"+ str(uncertain_input_sheet.cell(row=i+2, column=3).value)
            name_repository.append(sample_string)


        self.energyPlus_input_setup.update({'names': name_repository})
        
##        print(self.energyPlus_input_setup)
##        print(len(self.energyPlus_input_setup))

        
        
        # Uncertainty quantification (only parameter uncertainty)
        design_lhs = lhs(self.number_of_parameter_uncertain_parameters, samples=self.number_of_samples)
        value = 0
        for i in range(0,self.number_of_parameter_uncertain_parameters):
            if uncertain_input_sheet.cell(row=i+2, column=5).value == "NormalRelative":
                replace_to_EnergyPlus_format = str(uncertain_input_sheet.cell(row=i+2, column=2).value.replace("_", ":"))
                print(replace_to_EnergyPlus_format)
                
                obj_num = int(uncertain_input_sheet.cell(row=i+2, column=4).value[3])
##                print(obj_num)
                
                field_name= str(uncertain_input_sheet.cell(row=i+2, column=3).value)
##                print(field_name)
##                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                
                if replace_to_EnergyPlus_format == 'Material':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    print(replace_to_EnergyPlus_format)
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Conductivity':
                        value = value_in_idf_instance.Conductivity
                        print(value)
            
                    elif field_name == 'Thickness':
                        value = value_in_idf_instance.Thickness

                    elif field_name == 'Density':
                        value = value_in_idf_instance.Density

                    elif field_name == 'Specific_Heat':
                        value = value_in_idf_instance.Specific_Heat

                elif replace_to_EnergyPlus_format == 'WindowMaterial_SimpleGlazingSystem':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'U-Factor':
                        value = value_in_idf_instance.U-Factor

                    elif field_name == 'Solar_heat_Gain_Coefficient':
                        value =  value_in_idf_instance.Solar_heat_Gain_Coefficient            

                elif replace_to_EnergyPlus_format == 'People':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'People_per_Zone_Floor_Area':
                        value_in_idf_instance = value_in_idf_instance.People_per_Zone_Floor_Area

                    elif field_name == 'Zone_Floor_Area_per_Person':
                        value_in_idf_instance = value_in_idf_instance.Zone_Floor_Area_per_Person
            
                elif replace_to_EnergyPlus_format == 'Lights':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Lighting_Level':
                        value = value_in_idf_instance.Lighting_Level

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value = value_in_idf_instance.Watts_per_Zone_Floor_Area

                    elif field_name == 'Watts_per_Person':
                        value = value_in_idf_instance.Watts_per_Person

                elif replace_to_EnergyPlus_format == 'ElectricEquipment':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Design_Level':
                        value = value_in_idf_instance.Design_Level

                    elif field_name == 'Watts_per_Zone_Floor_Area':
                        value = value_in_idf_instance.Watts_per_Zone_Floor_Area

                    elif field_name == 'Watts_per_Person':
                        value = value_in_idf_instance.Watts_per_Person

                elif replace_to_EnergyPlus_format == 'ZoneInfiltration_DesignFlowRate':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Design_Flow_Rate':
                        value = value_in_idf_instance.Design_Flow_Rate

                    elif field_name == 'Flow_per_Zone_Floor_Area':
                        value = value_in_idf_instance.Flow_per_Zone_Floor_Area

                    elif field_name == 'Flow_per_Exterior_Surface_Area':
                        value = value_in_idf_instance.Flow_per_Exterior_Surface_Area

                    elif field_name == 'Air_Changes_per_Hour':
                        value = value_in_idf_instance.Air_Changes_per_Hour
                
                elif replace_to_EnergyPlus_format == 'DesignSpecification_OutdoorAir':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Outdoor_Air_Flow_per_Person':
                        value = value_in_idf_instance.Outdoor_Air_Flow_per_Person

                    elif field_name == 'Outdoor_Air_Flow_per_Zone_Floor_Area':
                        value = value_in_idf_instance.Outdoor_Air_Flow_per_Zone_Floor_Area

                    elif field_name == 'Outdoor_Air_Flow_per_Zone':
                        value = value_in_idf_instance.Outdoor_Air_Flow_per_Zone

                    elif field_name == 'Outdoor_Air_Flow_Air_Changes_per_Hour':
                        value = value_in_idf_instance.Outdoor_Air_Flow_Air_Changes_per_Hour

                elif replace_to_EnergyPlus_format == 'AirflowNetwork_MultiZone_Surface_EffectiveLeakageArea':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Effective_Leakage_Area':
                        value = value_in_idf_instance.Effective_Leakage_Area                   

                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Wind_Pressure_Coefficient_Value_1':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_1

                    elif field_name == 'Wind_Pressure_Coefficient_Value_2':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_2

                    elif field_name == 'Wind_Pressure_Coefficient_Value_3':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_3

                    elif field_name == 'Wind_Pressure_Coefficient_Value_4':
                        value.Wind_Pressure_Coefficient_Value_4

                    elif field_name == 'Wind_Pressure_Coefficient_Value_5':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_5

                    elif field_name == 'Wind_Pressure_Coefficient_Value_6':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_6

                    elif field_name == 'Wind_Pressure_Coefficient_Value_7':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_7

                    elif field_name == 'Wind_Pressure_Coefficient_Value_8':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_8

                    elif field_name == 'Wind_Pressure_Coefficient_Value_9':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_9

                    elif field_name == 'Wind_Pressure_Coefficient_Value_10':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_10

                    elif field_name == 'Wind_Pressure_Coefficient_Value_11':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_11

                    elif field_name == 'Wind_Pressure_Coefficient_Value_12':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_12

                    elif field_name == 'Wind_Pressure_Coefficient_Value_13':
                        value = value_in_idf_instance.Wind_Pressure_Coefficient_Value_13

                elif replace_to_EnergyPlus_format == 'Fan_ConstantVolume':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Pressure_Rise':
                        value = value_in_idf_instance.Pressure_Rise

                    elif field_name == 'Maximum_Flow_Rate':
                        value = value_in_idf_instance.Maximum_Flow_Rate

                elif replace_to_EnergyPlus_format == 'Fan_VariableVolume':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Fan_Total_Efficiency':
                        value = value_in_idf_instance.Fan_Total_Efficiency

                    elif field_name == 'Pressure_Rise':
                        value = value_in_idf_instance.Pressure_Rise

                elif replace_to_EnergyPlus_format == 'Coil_Cooling_DX_SingleSpeed':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value = value_in_idf_instance.Gross_Rated_Total_Cooling_Capacity

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value = value_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value = value_in_idf_instance.Gross_Rated_Cooling_COP

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value = value_in_idf_instance.Rated_Air_Flow_Rate
                                       
                elif replace_to_EnergyPlus_format == 'Coil_Cooling_DX_TwoSpeed':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'High_Speed_Gross_Rated_Total_Cooling_Capacity':
                        value = value_in_idf_instance.High_Speed_Gross_Rated_Total_Cooling_Capacity

                    elif field_name == 'High_Speed_Rated_Sensible_Heat_Ratio':
                        value = value_in_idf_instance.High_Speed_Rated_Sensible_Heat_Ratio

                    elif field_name == 'High_Speed_Gross_Rated_Cooling_COP':
                        value = value_in_idf_instance.High_Speed_Gross_Rated_Cooling_COP

                    elif field_name == 'High_Speed_Rated_Air_Flow_Rate':
                        value = value_in_idf_instance.High_Speed_Rated_Air_Flow_Rate

                elif replace_to_EnergyPlus_format == 'CoilPerformance_DX_Cooling':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Gross_Rated_Total_Cooling_Capacity':
                        value = value_in_idf_instance.Gross_Rated_Total_Cooling_Capacity

                    elif field_name == 'Gross_Rated_Sensible_Heat_Ratio':
                        value = value_in_idf_instance.Gross_Rated_Sensible_Heat_Ratio

                    elif field_name == 'Gross_Rated_Cooling_COP':
                        value = value_in_idf_instance.Gross_Rated_Cooling_COP

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value = value_in_idf_instance.Rated_Air_Flow_Rate

                elif replace_to_EnergyPlus_format == 'Coil_Heating_DX_SingleSpeed':
                    replace_to_EnergyPlus_format = replace_to_EnergyPlus_format.upper()
                    value_in_idf_instance = incetance_idf.idfobjects[replace_to_EnergyPlus_format][obj_num-1]
                    if field_name == 'Gross_Rated_Heating_Capacity':
                        value = value_in_idf_instance.Gross_Rated_Heating_Capacity

                    elif field_name == 'Gross_Rated_Heating_COP':
                        value = value_in_idf_instance.Gross_Rated_Heating_COP

                    elif field_name == 'Rated_Air_Flow_Rate':
                        value = value_in_idf_instance.Rated_Air_Flow_Rate

                self.distribution_repository[:,i] = norm(loc=value, scale=uncertain_input_sheet.cell(row=i+2, column=6).value).ppf(design_lhs[:,i])
                
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "UniformRelative":                
                self.distribution_repository[:,i] = uniform(loc=uncertain_input_sheet.cell(row=i+2, column=6).value, scale=uncertain_input_sheet.cell(row=i+2, column=7).value).ppf(design_lhs[:,i])
                
            elif uncertain_input_sheet.cell(row=i+2, column=5).value == "TriangleRelative":
                self.distribution_repository[:,i] = triang(c=uncertain_input_sheet.cell(row=i+2, column=6).value, loc=uncertain_input_sheet.cell(row=i+2, column=7).value, scale=uncertain_input_sheet.cell(row=i+2, column=8).value).ppf(design_lhs[:,i])

             
        incetance_idf.save() # close idf 
        uncertain_input.save('Uncertain_EP_Input.xlsx') # close Excel

##        print(self.distribution_repository)
        
        # Conduct Morris Method
        Y = self.EP_iteration(self.distribution_repository)
        Si = morris.analyze(self.energyPlus_input_setup, self.distribution_repository, Y, conf_level=0.95, print_to_console=True, num_levels=4)


        # Visualize the morris method result


##    def UA(self):
##
##        # Conduct uncertainty Analysis
##        for i in range(0,self.number_of_samples):
##            
##            # Open IDF file
##            IDF.setiddname(self.IDD_FileName)
##            idf1 = IDF(self.IDF_FileName)
##
##            # Assign quantified values in the idf instace
##            lightins.Watts_per_Zone_Floor_Area = X[0]
##
##
##            # Run
##            idf1.save()
##            idf = IDF(self.IDF_FileName, self.epw_FileName)
##            idf.run(readvars=True)

            
        # Visualize the Uncertainty Analysis Results
        

testing = Uncertain_EP('New_TGS.idf', 'Atlanta.epw', 'Energy+.idd')
testing.SA()
